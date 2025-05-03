# Bloque: Generar hoja 'Auditoría Sucursales' con formato validado
import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO

# Filtrar y procesar la base cargada
df = df_base.copy()

# Clasificación del riesgo
def clasificar_riesgo(monto):
    if monto >= 60000:
        return "🔴 Crítico"
    elif monto >= 30000:
        return "🟡 Moderado"
    else:
        return "🟢 Bajo"

df["Grupo de Riesgo"] = df.groupby("Categoría")["Monto"].transform("sum").apply(clasificar_riesgo)

# Calcular gasto total por sucursal
total_por_sucursal = df.groupby("Sucursales")["Monto"].transform("sum")
df["Gasto Total de Sucursal"] = total_por_sucursal
df["Participación (%)"] = (df["Monto"] / total_por_sucursal * 100).round(2)

# Columna ¿Prioridad para Revisión?
df["¿Prioridad para Revisión?"] = df.apply(
    lambda row: "✅ Sí" if row["Grupo de Riesgo"] == "🔴 Crítico" or row["Participación (%)"] > 30 else "🔍 No", axis=1
)

# Selección y orden de columnas
df_export = df[[
    "Sucursales", "Grupo de Riesgo", "Categoria", "Descripcion", "Fecha",
    "Monto", "Gasto Total de Sucursal", "Participación (%)", "¿Prioridad para Revisión?"
]].copy()

# Formato adicional
df_export["Monto"] = df_export["Monto"].apply(lambda x: f"{x:,.2f}")
df_export["Gasto Total de Sucursal"] = df_export["Gasto Total de Sucursal"].apply(lambda x: f"{x:,.2f}")
df_export["Fecha"] = pd.to_datetime(df_export["Fecha"]).dt.strftime('%Y-%m-%d')

# Ordenar por prioridad y participación
df_export.sort_values(by=["¿Prioridad para Revisión?", "Participación (%)"], ascending=[False, False], inplace=True)

# Generar archivo Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Auditoría Sucursales"

# Encabezado institucional
ws["A1"] = "Auditoría grupo Farmavalue"
ws["A1"].font = Font(size=28, bold=True, color="FF0000")
ws.merge_cells("A1:L1")

ws["A2"] = "Reporte de gastos del 01 de Enero al 20 de abril del 2025"
ws.merge_cells("A2:L2")

ws["A3"] = "Auditor Asignado:"
ws["A4"] = "Fecha de la Auditoría"

# Escribir encabezado de tabla desde A6
headers = list(df_export.columns) + ["Verificado", "No Verificado", "Comentario del Auditor"]
ws.append(headers)

# Estilo para encabezados
for col in range(1, len(headers)+1):
    cell = ws.cell(row=6, column=col)
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="D9D9D9")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Agregar datos desde fila 7
for row in df_export.itertuples(index=False):
    fila = list(row) + ["☐", "☐", ""]
    ws.append(fila)

# Ajustar alineaciones y justificación
for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(horizontal="left", vertical="center")

# Guardar archivo en memoria
output = BytesIO()
wb.save(output)
output.seek(0)

# Botón de descarga
st.markdown("### 📥 Descargar reporte de Auditoría por Sucursales")
st.download_button(
    label="📊 Descargar Auditoría Sucursales",
    data=output,
    file_name="Auditoria_Sucursales_Formato_Final_OK_v2.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
